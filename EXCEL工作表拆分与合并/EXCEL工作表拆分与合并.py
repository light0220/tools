import os
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.utils import get_column_letter


class SheetCopy:
    def __init__(self) -> None:
        pass

    # 定义工作表复制模块
    def copy_sheet(self, src_file_sheet, tag_file_sheet):
        for row in src_file_sheet:
            # 遍历源xlsx文件制定sheet中的所有单元格
            for cell in row:  # 复制数据
                tag_file_sheet[cell.coordinate].value = cell.value
                if cell.has_style:  # 复制样式
                    tag_file_sheet[cell.coordinate].font = copy(cell.font)
                    tag_file_sheet[cell.coordinate].border = copy(cell.border)
                    tag_file_sheet[cell.coordinate].fill = copy(cell.fill)
                    tag_file_sheet[cell.coordinate].number_format = copy(
                        cell.number_format)
                    tag_file_sheet[cell.coordinate].protection = copy(
                        cell.protection)
                    tag_file_sheet[cell.coordinate].alignment = copy(
                        cell.alignment)

        wm = list(zip(src_file_sheet.merged_cells))  # 开始处理合并单元格
        if len(wm) > 0:  # 检测源xlsx中合并的单元格
            for i in range(0, len(wm)):
                cell2 = (str(wm[i]).replace(
                    "(<MergedCellRange ", "").replace(">,)", ""))  # 获取合并单元格的范围
                tag_file_sheet.merge_cells(cell2)  # 合并单元格
        # 开始处理行高列宽
        for i in range(1, src_file_sheet.max_row + 1):
            tag_file_sheet.row_dimensions[i].height = src_file_sheet.row_dimensions[i].height
        for i in range(1, src_file_sheet.max_column + 1):
            tag_file_sheet.column_dimensions[get_column_letter(
                i)].width = src_file_sheet.column_dimensions[get_column_letter(i)].width + self.column_adjust  # 修正列宽误差

    # 定义工作薄拆分模块
    def split_workbook(self):
        layout1 = [
            [sg.In(key='-PATH-'), sg.FileBrowse(button_text='浏览···',
                                                file_types=[['Excel文件', '*.xlsx']])]
        ]
        layout2 = [
            [sg.In(key='-FOLDER-'), sg.FolderBrowse(button_text='浏览···')]
        ]
        layout3 = [
            [sg.R('将所有文件存放在同一目录下', key='-TOG-', group_id=0, default=True)],
            [sg.R('将文件存放在按工作表命名的不同目录下', key='-SEP-', group_id=0)]
        ]
        layout = [
            [sg.Frame('请选择要拆分的Excel文件', layout1, size=(400, 60))],
            [sg.Frame('请选择拆分后的文件输出目录', layout2, size=(400, 60))],
            [sg.Frame('请选择拆分后文件存放方式', layout3, size=(400, 80))],
            [sg.Button('确定'), sg.Button('取消')]
        ]
        window = sg.Window('Excel工作薄拆分', layout)
        self.column_adjust = -0.09  # 修正列宽误差
        while True:
            event, values = window.read()
            if event == None:
                break
            if event == '取消':
                break
            if event == '确定':
                if len(values['-PATH-']) * len(values['-FOLDER-']) == 0:
                    sg.popup('请选择要拆分的文件及输出目录')
                    continue
                self.source_file = values['-PATH-']
                self.object_path = values['-FOLDER-']
                src_workbook = load_workbook(self.source_file)  # 打开源xlsx
                for sheet_name in src_workbook.sheetnames:
                    src_file_sheet = src_workbook[sheet_name]  # 打开目标sheet
                    tag_workbook = Workbook()  # 新建空工作薄
                    tag_file_sheet = tag_workbook.active
                    tag_file_sheet.title = sheet_name
                    self.copy_sheet(src_file_sheet, tag_file_sheet)
                    if values['-TOG-'] == True:
                        tag_workbook.save(
                            f'{self.object_path}/{sheet_name}.xlsx')
                    if values['-SEP-'] == True:
                        os.makedirs(
                            f'{self.object_path}/{sheet_name}', exist_ok=True)
                        tag_workbook.save(
                            f'{self.object_path}/{sheet_name}/{sheet_name}.xlsx')
                sg.popup('工作薄拆分完成')
                break
        window.close()

    # 定义工作表合并模块
    def merge_sheets(self):
        layout1 = [
            [sg.In(key='-FOLDER-'), sg.FolderBrowse(button_text='浏览···')]
        ]
        layout2 = [
            [sg.In(key='-SAVEAS-'), sg.SaveAs(button_text='浏览···',
                                              file_types=[['Excel文件', '*.xlsx']])]
        ]
        layout = [
            [sg.Frame('请将要合并的Excel文件放入同一目录并选择该目录', layout1, size=(400, 60))],
            [sg.Frame('请选择合并后文件保存路径', layout2, size=(400, 60))],
            [sg.Button('确定'), sg.Button('取消')]
        ]
        window = sg.Window('Excel工作表合并', layout)
        self.column_adjust = 0  # 修正列宽误差
        while True:
            event, values = window.read()
            if event == None:
                break
            if event == '取消':
                break
            if event == '确定':
                if len(values['-SAVEAS-']) * len(values['-FOLDER-']) == 0:
                    sg.popup('请选择要合并的文件目录及合并后的文件保存路径')
                    continue
                self.source_file = values['-FOLDER-']
                self.object_path = values['-SAVEAS-']
                tag_workbook = Workbook()  # 新建空工作薄
                ws = tag_workbook.active
                tag_workbook.remove(ws)  # 删除默认工作表
                for filename in os.listdir(self.source_file):
                    src_workbook = load_workbook(
                        f'{self.source_file}/{filename}')  # 打开源xlsx
                    for sheet_name in src_workbook.sheetnames:
                        src_file_sheet = src_workbook[sheet_name]  # 打开目标sheet
                        tag_file_sheet = tag_workbook.create_sheet(sheet_name)
                        self.copy_sheet(src_file_sheet, tag_file_sheet)
                tag_workbook.save(self.object_path)
                sg.popup('工作表合并完成')
                break
        window.close()

    # 主程序
    def main(self):
        sg.theme('DefaultNoMoreNagging')
        layout1 = [
            [sg.B('工作薄拆分', size=(
                20, 2), tooltip='将一个Excel文件中的多个工作表拆分为多个Excel文件', pad=20, font=('微软雅黑', 20))],
            [sg.B('工作表合并', size=(
                20, 2), tooltip='将同目录下的多个Excel文件中的工作表合并为一个Excel文件', pad=20, font=('微软雅黑', 20))],
            [sg.B('退出', size=(20, 2), tooltip='退出', pad=20, font=('微软雅黑', 20))]
        ]
        layout = [
            [sg.Frame('请选择功能', layout1)],
            [sg.T('by:', font=('黑体', 12), text_color='red', size=28, justification='right'), sg.T(
                '北极星光', font=('黑体', 12), text_color='red')],
            [sg.T('E-mail:', font=('黑体', 12), text_color='red', size=28, justification='right'),
             sg.T('light22@126.com', font=('黑体', 12), text_color='red')]
        ]
        window = sg.Window('Excel小工具', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            if event == '退出':
                break
            if event == '工作薄拆分':
                self.split_workbook()
            if event == '工作表合并':
                self.merge_sheets()
        window.close()


# 调试
if __name__ == '__main__':
    sg.popup(
        f'''{'='*20}欢迎使用本小工具{'='*20}''',
        '''1.本工具为Excel文件工作表拆分及合并工具，可将单个Excel文
  件中的多个工作表拆分为多个工作表，也可将多个Excel文件
  中的工作表合并为一个Excel文件。
2.本工具支持扩展名为.xlsx的文件，不支持Excel 97~2003（.x
  ls）文件格式,如需转换请在Excel中另存为.xlsx文件再使用
  本工具转换。
3.本工具可以复制大多数Excel文件的数据、样式、单元格格式
  等内容，无法复制批注内容，部分由专业软件生成的Excel文
  件可能样式会复制不全，但数据基本上会复制全。
4.由于计算机运算机制差异，部分列宽可能会存在1以内的误差。
5.本工具可自由传播，但禁止用于商业用途。
6.BUG反馈及其它未尽事宜请致电:light22@126.com。
''', text_color='yellow', background_color='green', font=('宋体', 15), no_titlebar=True, any_key_closes=True
    )
    SheetCopy().main()
