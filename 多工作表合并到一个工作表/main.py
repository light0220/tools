#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
作  者    : 北极星光 light22@126.com
创建时间  : 2024-01-02 15:08:41
最后编辑  : 2024-01-02 16:26:08
编辑人    : 北极星光
'''

import os
from openpyxl import Workbook,load_workbook

# 指定工作路径
path = r'E:\Codes\tools\多工作表合并到一个工作表\示例表格'

# 新建一个空工作表用来存放数据
new_wb = Workbook()
new_ws = new_wb.active

# 循环得到所有工作表
for filename in os.listdir(path):
    wb = load_workbook(f'{path}\\{filename}')
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # 通过姓名列取值循环获取当前工作表最大行
        max_row = 15  # 初始从第15行开始
        while ws[f'C{max_row}'].value:
            max_row += 1
        for row in ws.iter_rows(min_row = 15, max_row = max_row-1):
            l = [i.value for i in row]
            new_ws.append(l)

# 保存工作表
new_wb.save(f'{path}\\合并后.xlsx')